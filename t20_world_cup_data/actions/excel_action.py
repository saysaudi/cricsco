from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

class ExcelAction:
    def __init__(self):
        self.workbook = Workbook()
        print("Workbook created successfully.")

    def write_match_details(self, match_details_list):
        scorecard_sheet = self.workbook.active
        scorecard_sheet.title = "Scorecard"
        for index in range(len(match_details_list)):
            current_cell = scorecard_sheet.cell(row=index+1, column=1, value=match_details_list[index])
            current_cell.font = Font(bold=True)

        self.workbook.save("cricket_data.xlsx")
        return True

    def write_batting_stats(self, bat_stat_list):
        self.bat_stat_list = bat_stat_list
        scorecard_sheet = self.workbook.active
        last_used_row = self.get_last_used_row()
        # Change the background color to yellow
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        count = 0
        for bat_stat in bat_stat_list:
            if count == 0: 
                for i in range(len(bat_stat)):
                    current_cell = scorecard_sheet.cell(row=last_used_row+1, column=i+1, value=bat_stat[i])
                    current_cell.font = Font(bold=True)
                    current_cell.fill = yellow_fill
                last_used_row = self.get_last_used_row()
            else:
                for i in range(len(bat_stat)):
                    scorecard_sheet.cell(row=last_used_row+i+1, column=count, value=bat_stat[i])
            count += 1
            self.workbook.save("cricket_data.xlsx")

    def get_last_used_row(self):
        scorecard_sheet = self.workbook.active
        # Iterate from the last row to the first row
        for row in reversed(range(1, scorecard_sheet.max_row + 1)):
            for col in scorecard_sheet.iter_cols(1, scorecard_sheet.max_column):
                if col[row-1].value is not None:
                    return row
